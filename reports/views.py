from datetime import date, timedelta
from io import BytesIO
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
from django.contrib import messages
from core.permissions import group_required
from .models import ReportTemplate, ScheduledReport, ReportExport
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def _get_report_data(category, branch_id=None, start_date=None, end_date=None):
    from billing.models import Bill, Payment
    from operations.models import Expense
    from orders.models import Order

    if not end_date:
        end_date = timezone.now().date()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    day_start = timezone.make_aware(
        timezone.datetime.combine(start_date, timezone.datetime.min.time())
    )
    day_end = timezone.make_aware(
        timezone.datetime.combine(end_date, timezone.datetime.max.time())
    )

    bills = Bill.objects.filter(generated_at__gte=day_start, generated_at__lte=day_end)
    payments = Payment.objects.filter(created_at__gte=day_start, created_at__lte=day_end)
    expenses = Expense.objects.filter(expense_date__gte=start_date, expense_date__lte=end_date)
    orders = Order.objects.filter(submitted_at__gte=day_start, submitted_at__lte=day_end)

    if branch_id:
        bills = bills.filter(branch_id=branch_id)
        payments = payments.filter(bill__branch_id=branch_id)
        expenses = expenses.filter(branch_id=branch_id)
        orders = orders.filter(session__table__floor__branch_id=branch_id)

    result = {"start_date": start_date, "end_date": end_date}

    if category == "SALES":
        result["total_sales"] = bills.aggregate(s=Sum("grand_total"))["s"] or Decimal("0")
        result["bill_count"] = bills.count()
        cash = payments.filter(payment_method="CASH").aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        card = payments.filter(payment_method="CARD").aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        online = payments.exclude(payment_method__in=["CASH", "CARD"]).aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        result["total_cash"] = cash
        result["total_card"] = card
        result["total_online"] = online
        result["order_count"] = orders.count()
        result["avg_bill"] = (result["total_sales"] / result["bill_count"]) if result["bill_count"] > 0 else Decimal("0")
        bills_by_date = bills.annotate(
            day=TruncDate("generated_at")
        ).values("day").annotate(
            total=Sum("grand_total"), cnt=Count("id")
        ).order_by("day")
        result["daily_breakdown"] = list(bills_by_date)

    elif category == "FINANCIAL":
        result["total_sales"] = bills.aggregate(s=Sum("grand_total"))["s"] or Decimal("0")
        result["total_expenses"] = expenses.aggregate(s=Sum("amount"))["s"] or Decimal("0")
        result["total_cash"] = payments.filter(payment_method="CASH").aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        result["total_card"] = payments.filter(payment_method="CARD").aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        result["total_online"] = payments.exclude(payment_method__in=["CASH", "CARD"]).aggregate(s=Sum("amount_paid"))["s"] or Decimal("0")
        result["net_revenue"] = result["total_sales"] - result["total_expenses"]
        result["bill_count"] = bills.count()

    elif category == "INVENTORY":
        from inventory.models import StockMovement, InventoryItem
        items = InventoryItem.objects.filter(is_active=True)
        if branch_id:
            items = items.filter(branch_id=branch_id)
        result["items"] = list(items.values("name", "quantity_in_stock", "low_stock_threshold", "unit_cost"))
        movements = StockMovement.objects.filter(created_at__gte=day_start, created_at__lte=day_end)
        if branch_id:
            movements = movements.filter(item__branch_id=branch_id)
        result["total_in"] = movements.filter(movement_type="IN").aggregate(s=Sum("quantity"))["s"] or 0
        result["total_out"] = movements.filter(movement_type="OUT").aggregate(s=Sum("quantity"))["s"] or 0

    elif category == "STAFF":
        from staff.models import TimeEntry, StaffProfile
        entries = TimeEntry.objects.filter(clock_in__gte=day_start, clock_in__lte=day_end)
        if branch_id:
            entries = entries.filter(staff__branch_id=branch_id)
        result["total_hours"] = entries.aggregate(s=Sum("total_hours"))["s"] or Decimal("0")
        result["total_overtime"] = entries.aggregate(s=Sum("overtime_hours"))["s"] or Decimal("0")
        result["staff_count"] = StaffProfile.objects.filter(status="ACTIVE").count()
        result["entries"] = list(entries.select_related("staff__user").values(
            "staff__user__first_name", "staff__user__last_name",
            "clock_in", "clock_out", "total_hours", "overtime_hours"
        ).order_by("-clock_in")[:100])

    elif category == "OPERATIONS":
        from operations.models import CashRegister, DayEndSummary
        registers = CashRegister.objects.filter(opened_at__gte=day_start, opened_at__lte=day_end)
        if branch_id:
            registers = registers.filter(branch_id=branch_id)
        result["register_count"] = registers.count()
        summaries = DayEndSummary.objects.filter(summary_date__gte=start_date, summary_date__lte=end_date)
        if branch_id:
            summaries = summaries.filter(branch_id=branch_id)
        result["day_end_summaries"] = list(summaries.select_related("branch").values(
            "branch__name", "summary_date", "total_sales", "total_expenses",
            "net_revenue", "is_closed"
        ).order_by("-summary_date"))

    return result


def _generate_excel(report_data, template, params):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template.name[:31]
    bold = openpyxl.styles.Font(bold=True)
    ws.append([f"{template.name} Report"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws.append([f"Period: {params.get('start_date', '')} to {params.get('end_date', '')}"])
    ws.append([])

    if isinstance(report_data, dict):
        for key, value in report_data.items():
            if key in ("start_date", "end_date", "daily_breakdown", "day_end_summaries", "items", "entries"):
                continue
            ws.append([key.replace("_", " ").title(), str(value)])
        ws.append([])

        for section_name in ["daily_breakdown", "day_end_summaries", "items", "entries"]:
            if section_name in report_data and report_data[section_name]:
                ws.append([section_name.replace("_", " ").title()])
                rows = report_data[section_name]
                if rows:
                    headers = list(rows[0].keys())
                    ws.append([h.replace("_", " ").title() for h in headers])
                    for row in rows:
                        ws.append([str(row.get(h, "")) for h in headers])
                    ws.append([])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_pdf(report_data, template, params):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16, spaceAfter=6)
    heading_style = ParagraphStyle("Heading2", parent=styles["Heading2"], fontSize=12, spaceAfter=4)
    normal_style = styles["Normal"]
    elements = []

    elements.append(Paragraph(f"{template.name} Report", title_style))
    elements.append(Paragraph(f"Period: {params.get('start_date', '')} to {params.get('end_date', '')}", normal_style))
    elements.append(Spacer(1, 10*mm))

    summary_data = [["Metric", "Value"]]
    for key, value in report_data.items():
        if key in ("start_date", "end_date", "daily_breakdown", "day_end_summaries", "items", "entries"):
            continue
        if isinstance(value, Decimal):
            display_val = f"Rs. {value:,.2f}"
        elif isinstance(value, float):
            display_val = f"Rs. {value:,.2f}"
        else:
            display_val = str(value)[:80]
        summary_data.append([key.replace("_", " ").title(), display_val])

    if len(summary_data) > 1:
        tbl = Table(summary_data, colWidths=[200, 280])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 10*mm))

    for section_name in ["daily_breakdown", "day_end_summaries", "items", "entries"]:
        section_data = report_data.get(section_name)
        if section_data:
            elements.append(Paragraph(section_name.replace("_", " ").title(), heading_style))
            headers = [h.replace("_", " ").title() for h in section_data[0].keys()]
            rows = [headers]
            for row in section_data:
                rows.append([str(row.get(h, ""))[:40] for h in section_data[0].keys()])
            tbl2 = Table(rows)
            tbl2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(tbl2)
            elements.append(Spacer(1, 8*mm))

    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(f"Generated by CafeFlow on {timezone.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
    doc.build(elements)
    buf.seek(0)
    return buf


@group_required("Admin", "Manager")
def report_list(request):
    templates = ReportTemplate.objects.filter(is_active=True)
    category = request.GET.get("category")
    if category:
        templates = templates.filter(category=category)
    return render(request, "reports/report_list.html", {"templates": templates})


@group_required("Admin", "Manager")
def report_view(request, template_id):
    template = get_object_or_404(ReportTemplate, pk=template_id)
    from organisation.models import Branch
    branches = Branch.objects.filter(is_active=True)
    return render(request, "reports/report_view.html", {
        "template": template,
        "branches": branches,
        "today": date.today().isoformat(),
    })


@group_required("Admin", "Manager")
def download_report(request, template_id):
    template = get_object_or_404(ReportTemplate, pk=template_id)
    fmt = request.GET.get("format", template.default_format)
    branch_id = request.GET.get("branch") or None
    period = request.GET.get("period", "monthly")
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    end_d = date.today()
    if period == "daily":
        start_d = end_d
    elif period == "weekly":
        start_d = end_d - timedelta(days=7)
    else:
        start_d = end_d - timedelta(days=30)

    if start_str:
        try:
            start_d = date.fromisoformat(start_str)
        except (ValueError, TypeError):
            pass
    if end_str:
        try:
            end_d = date.fromisoformat(end_str)
        except (ValueError, TypeError):
            pass

    report_data = _get_report_data(template.category, branch_id=branch_id, start_date=start_d, end_date=end_d)

    if fmt == "EXCEL":
        buf = _generate_excel(report_data, template, {"start_date": start_d, "end_date": end_d})
        filename = f"{template.name}_{start_d}_{end_d}.xlsx"
        resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        ReportExport.objects.create(report=template, generated_by=request.user, format="EXCEL", parameters={
            "branch_id": str(branch_id) if branch_id else None,
            "start_date": str(start_d), "end_date": str(end_d),
        })
        return resp

    if fmt == "PDF":
        buf = _generate_pdf(report_data, template, {"start_date": start_d, "end_date": end_d})
        filename = f"{template.name}_{start_d}_{end_d}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        ReportExport.objects.create(report=template, generated_by=request.user, format="PDF", parameters={
            "branch_id": str(branch_id) if branch_id else None,
            "start_date": str(start_d), "end_date": str(end_d),
        })
        return resp

    if fmt == "HTML":
        return render(request, "reports/report_data.html", {
            "template": template, "data": report_data,
            "start_date": start_d, "end_date": end_d,
        })

    return render(request, "reports/report_data.html", {
        "template": template, "data": report_data,
        "start_date": start_d, "end_date": end_d,
    })


@group_required("Admin", "Manager")
def scheduled_reports(request):
    schedules = ScheduledReport.objects.select_related("report").all()
    return render(request, "reports/scheduled.html", {"schedules": schedules})


@group_required("Admin", "Manager")
def quick_report(request):
    category = request.GET.get("category", "SALES")
    fmt = request.GET.get("format", "EXCEL")
    period = request.GET.get("period", "daily")
    branch_id = request.GET.get("branch") or None

    end_d = date.today()
    if period == "daily":
        start_d = end_d
    elif period == "weekly":
        start_d = end_d - timedelta(days=7)
    else:
        start_d = end_d - timedelta(days=30)

    report_data = _get_report_data(category, branch_id=branch_id, start_date=start_d, end_date=end_d)

    templates = ReportTemplate.objects.filter(category=category, is_active=True)
    template_name = templates.first().name if templates.exists() else f"{category} Report"

    params = {"start_date": start_d, "end_date": end_d}
    if fmt == "PDF":
        buf = _generate_pdf(report_data, ReportTemplate(name=template_name, category=category), params)
        filename = f"{template_name}_{period}_{start_d}_{end_d}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    buf = _generate_excel(report_data, ReportTemplate(name=template_name, category=category), params)
    filename = f"{template_name}_{period}_{start_d}_{end_d}.xlsx"
    resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
